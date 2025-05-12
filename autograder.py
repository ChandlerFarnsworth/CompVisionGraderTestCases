#!/usr/bin/python3
"""
Excel Worksheet Autograder for Coursera

This script handles submission processing and grading for Excel worksheets.
Includes hidden test cases to verify specific cells.
"""

import os
import sys
import shutil
import json
import openpyxl
from pathlib import Path

# Constants (modify these for your assignment)
SUBMISSION_LOCATION = "/shared/submission"
SUBMISSION_DESTINATION = "/grader/submission.xlsx"
REFERENCE_SOLUTION = "/grader/solution.xlsx"
COURSERA_PARTID = "Lg9eS"  # Update with your assignment's part ID

# Worksheet names
STUDENT_SHEET_NAME = "blank"
SOLUTION_SHEET_NAME = "solution"

# Hidden test cells to check
HIDDEN_TEST_CELLS = [
    {"cell": "AD21", "description": "Financial calculation test"},
    {"cell": "M62", "description": "Data processing test"},
    {"cell": "AE187", "description": "Formula application test"}
]

def print_stderr(error_msg):
    """Print error message to stderr"""
    print(str(error_msg), file=sys.stderr)

def send_feedback(score, msg):
    """Send feedback to Coursera autograder"""
    post = {'fractionalScore': score, 'feedback': msg}
    print(json.dumps(post))
    
    # Write feedback to file for Coursera
    try:
        with open("/shared/feedback.json", "w") as outfile:
            json.dump(post, outfile)
    except Exception as e:
        print_stderr(f"Error writing feedback: {e}")

def grade_excel_worksheet():
    """
    Grade the Excel worksheet by comparing Y/N values in row 1
    and checking hidden test cells
    
    Returns:
        Dictionary with score and feedback
    """
    try:
        # Load workbooks
        student_wb = openpyxl.load_workbook(SUBMISSION_DESTINATION, data_only=True)
        solution_wb = openpyxl.load_workbook(REFERENCE_SOLUTION, data_only=True)
        
        # Verify sheets exist
        if STUDENT_SHEET_NAME not in student_wb.sheetnames:
            return {
                "score": 0.0,
                "feedback": f"Error: Worksheet '{STUDENT_SHEET_NAME}' not found in your submission."
            }
        
        if SOLUTION_SHEET_NAME not in solution_wb.sheetnames:
            return {
                "score": 0.0,
                "feedback": f"Error: Internal error - Solution worksheet not found."
            }
        
        # Get sheets
        student_sheet = student_wb[STUDENT_SHEET_NAME]
        solution_sheet = solution_wb[SOLUTION_SHEET_NAME]
        
        # PART 1: Check Y/N values in row 1
        matches = 0
        total_cells = 0
        max_col = max(solution_sheet.max_column, student_sheet.max_column)
        
        # Start from column E (index 5 in openpyxl)
        for col_idx in range(5, max_col + 1):
            student_cell = student_sheet.cell(row=1, column=col_idx)
            solution_cell = solution_sheet.cell(row=1, column=col_idx)
            
            if student_cell.value is not None and solution_cell.value is not None:
                total_cells += 1
                if student_cell.value == solution_cell.value:
                    matches += 1
        
        # PART 2: Check hidden test cells
        hidden_tests_passed = 0
        for test in HIDDEN_TEST_CELLS:
            cell_addr = test["cell"]
            
            # Get values from both sheets
            student_value = get_cell_value(student_sheet, cell_addr)
            solution_value = get_cell_value(solution_sheet, cell_addr)
            
            # Compare values (considering None and empty string as equal)
            if equal_values(student_value, solution_value):
                hidden_tests_passed += 1
        
        # Calculate total score (80% from Y/N values, 20% from hidden tests)
        yn_score = matches / total_cells if total_cells > 0 else 0.0
        hidden_score = hidden_tests_passed / len(HIDDEN_TEST_CELLS)
        
        total_score = yn_score * 0.8 + hidden_score * 0.2
        percentage = total_score * 100
        
        # Generate feedback (only showing Y/N results to the student)
        feedback = f"Your score: {percentage:.2f}%\nYou correctly matched {matches} out of {total_cells} cells."
        
        return {
            "score": total_score,
            "feedback": feedback
        }
        
    except Exception as e:
        return {
            "score": 0.0,
            "feedback": f"Error grading your submission: {str(e)}"
        }

def get_cell_value(sheet, cell_addr):
    """
    Safely get a cell value
    
    Args:
        sheet: The worksheet
        cell_addr: Cell address (e.g., "A1")
        
    Returns:
        Cell value or None if cell doesn't exist
    """
    try:
        return sheet[cell_addr].value
    except:
        return None

def equal_values(val1, val2):
    """
    Compare two cell values, treating None and empty string as equal
    Also handles numeric comparisons with small floating-point differences
    
    Args:
        val1: First value
        val2: Second value
        
    Returns:
        True if values are equal, False otherwise
    """
    # Both are None or empty
    if (val1 is None or val1 == "") and (val2 is None or val2 == ""):
        return True
    
    # Both are numeric
    if isinstance(val1, (int, float)) and isinstance(val2, (int, float)):
        # Allow small floating point differences
        return abs(val1 - val2) < 0.0001
    
    # String comparison (case-insensitive for text)
    if isinstance(val1, str) and isinstance(val2, str):
        return val1.strip().lower() == val2.strip().lower()
    
    # Direct comparison for other types
    return val1 == val2

def main(part_id):
    """Main function for the autograder"""
    # Verify correct part ID
    if part_id != COURSERA_PARTID:
        print_stderr("Cannot find matching partId. Please double check your partId's")
        send_feedback(0.0, "Please verify that you have submitted to the proper part of the assignment.")
        return
    
    # Find student submission
    learner_file = None
    for f in os.listdir(SUBMISSION_LOCATION):
        extension = Path(f).suffix.lower()
        if extension in ['.xlsx', '.xlsm']:
            learner_file = f
            break
    
    # Check if submission was found
    if learner_file is None:
        send_feedback(0.0, "Your submission file does not have the right file extension. Please submit an Excel file (.xlsx, .xlsm).")
        return
    
    # Copy submission to destination
    try:
        shutil.copyfile(os.path.join(SUBMISSION_LOCATION, learner_file), SUBMISSION_DESTINATION)
    except Exception as e:
        print_stderr(f"Error copying submission: {e}")
        send_feedback(0.0, "Error processing your submission file.")
        return
    
    # Grade submission
    result = grade_excel_worksheet()
    
    # Send feedback
    send_feedback(result["score"], result["feedback"])

if __name__ == "__main__":
    # Print Python version info
    print(sys.version_info)
    
    try:
        # Get part ID from environment
        part_id = os.environ.get('partId', '')
        
        # Run main function
        main(part_id)
    except Exception as e:
        print_stderr(f"Error in autograder: {e}")
        send_feedback(0.0, "Please provide the partId.")